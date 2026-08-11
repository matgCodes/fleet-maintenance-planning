import openpyxl
import json
import os
from collections import defaultdict

# Path to the primary Excel inventory
EXCEL_PATH = "/Users/mag_station/Library/CloudStorage/OneDrive2-PublicHealthInstitute/Asset_Tire_Inventory.xlsx"
TASKS_OUTPUT_PATH = "research_tasks.json"
MAPPING_OUTPUT_PATH = "task_asset_mapping.json"

def extract_tasks():
    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"Inventory Excel file not found at: {EXCEL_PATH}")

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active

    # Group asset IDs by unique vehicle key (Year | Make | Model)
    vehicle_groups = defaultdict(list)
    total_assets = 0

    for r in range(2, ws.max_row + 1):
        asset_id = ws.cell(row=r, column=1).value
        year = ws.cell(row=r, column=2).value
        make = ws.cell(row=r, column=3).value
        model = ws.cell(row=r, column=4).value

        if asset_id and year and make and model:
            asset_str = str(asset_id).strip()
            y_str = str(year).strip()
            m_str = str(make).strip().upper()
            mo_str = str(model).strip().upper()

            key = (y_str, m_str, mo_str)
            vehicle_groups[key].append({
                "row": r,
                "assetId": asset_str
            })
            total_assets += 1

    # Sort vehicle groups by frequency (most common first)
    sorted_groups = sorted(vehicle_groups.items(), key=lambda item: len(item[1]), reverse=True)

    tasks = []
    asset_mapping = {}

    for idx, (vehicle_tuple, assets) in enumerate(sorted_groups, start=1):
        task_id = f"veh_{idx:03d}"
        year, make, model = vehicle_tuple

        task_payload = {
            "taskId": task_id,
            "vehicle": {
                "year": year,
                "make": make,
                "model": model
            },
            "instructions": "Find the standard OEM factory tire size for the base/fleet model of this vehicle. Return the data adhering strictly to the response schema."
        }
        tasks.append(task_payload)

        asset_mapping[task_id] = {
            "vehicle": {
                "year": year,
                "make": make,
                "model": model
            },
            "count": len(assets),
            "assets": assets
        }

    # Write research_tasks.json
    with open(TASKS_OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(tasks, f, indent=2)

    # Write task_asset_mapping.json
    with open(MAPPING_OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(asset_mapping, f, indent=2)

    print(f"✅ Extraction Complete!")
    print(f"  - Total Fleet Assets Processed: {total_assets}")
    print(f"  - Unique Tasks Created: {len(tasks)}")
    print(f"  - Output Task File: {TASKS_OUTPUT_PATH}")
    print(f"  - Output Mapping File: {MAPPING_OUTPUT_PATH}")

if __name__ == "__main__":
    extract_tasks()
