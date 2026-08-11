#!/usr/bin/env python3
"""Create a new enriched tire workbook from complete Codex research results.

The source workbook is opened read-only from the user's perspective: this
script always saves to a distinct output path and verifies that the source file
hash did not change.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import tempfile
from collections import Counter
from pathlib import Path

import openpyxl
from openpyxl.styles import PatternFill

from codex_research_queue import ValidationError, load_json, task_index, validate_result_list


DEFAULT_SOURCE = Path(
    "/Users/mag_station/Library/CloudStorage/OneDrive2-PublicHealthInstitute/"
    "Asset_Tire_Inventory.xlsx"
)
DEFAULT_MAPPING = Path("task_asset_mapping.json")
DEFAULT_TASKS = Path("research_tasks.json")
DEFAULT_RESULTS = Path("codex_research_results.json")
DEFAULT_OUTPUT = Path("outputs/codex-tire-research/Asset_Tire_Inventory_Codex_Enriched.xlsx")
DEFAULT_AUDIT = Path("outputs/codex-tire-research/codex_research_audit.json")

FILL_HIGH = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")
FILL_MEDIUM = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
FILL_LOW = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")


def file_hash(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def fill_for_score(score: int) -> PatternFill:
    if score >= 4:
        return FILL_HIGH
    if score == 3:
        return FILL_MEDIUM
    return FILL_LOW


def validate_mapping(mapping: object, known_tasks: dict[str, dict]) -> dict[str, dict]:
    if not isinstance(mapping, dict):
        raise ValidationError("Task asset mapping must be a JSON object")
    missing = sorted(set(known_tasks) - set(mapping))
    extra = sorted(set(mapping) - set(known_tasks))
    if missing or extra:
        raise ValidationError(
            f"Mapping/task mismatch; missing={missing or 'none'}, extra={extra or 'none'}"
        )

    rows: list[int] = []
    for task_id, entry in mapping.items():
        assets = entry.get("assets") if isinstance(entry, dict) else None
        if not isinstance(assets, list) or not assets:
            raise ValidationError(f"{task_id}: mapping has no asset rows")
        for asset in assets:
            row = asset.get("row") if isinstance(asset, dict) else None
            if isinstance(row, bool) or not isinstance(row, int) or row < 2:
                raise ValidationError(f"{task_id}: invalid worksheet row {row!r}")
            rows.append(row)

    duplicates = sorted(row for row, count in Counter(rows).items() if count > 1)
    if duplicates:
        raise ValidationError(f"Worksheet rows are mapped more than once: {duplicates}")
    return mapping


def save_workbook_atomically(workbook, output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    fd, temporary_name = tempfile.mkstemp(
        dir=output_path.parent, prefix=f".{output_path.stem}.", suffix=".xlsx"
    )
    os.close(fd)
    try:
        workbook.save(temporary_name)
        os.replace(temporary_name, output_path)
    except Exception:
        try:
            os.unlink(temporary_name)
        except FileNotFoundError:
            pass
        raise


def inject(
    source: Path,
    output: Path,
    tasks_path: Path,
    mapping_path: Path,
    results_path: Path,
    audit_path: Path,
) -> tuple[int, int]:
    if source.resolve() == output.resolve():
        raise ValidationError("Output path must differ from the source workbook path")
    if output.exists():
        raise ValidationError(f"Output already exists; refusing to overwrite: {output}")
    if audit_path.exists():
        raise ValidationError(f"Audit output already exists; refusing to overwrite: {audit_path}")
    if not source.exists():
        raise ValidationError(f"Source workbook not found: {source}")

    tasks, known_tasks = task_index(tasks_path)
    mapping = validate_mapping(load_json(mapping_path), known_tasks)
    results = validate_result_list(load_json(results_path), known_tasks, str(results_path))
    result_by_id = {result["taskId"]: result for result in results}
    missing = [task["taskId"] for task in tasks if task["taskId"] not in result_by_id]
    if missing:
        raise ValidationError(
            f"Research is incomplete ({len(results)}/{len(tasks)}); first missing task: {missing[0]}"
        )

    source_hash_before = file_hash(source)
    workbook = openpyxl.load_workbook(source)
    worksheet = workbook.active
    expected_headers = [
        "Asset ID Number",
        "Year",
        "Make",
        "Model",
        "Front Tire Size",
        "Rear Tire Size",
    ]
    actual_headers = [worksheet.cell(1, column).value for column in range(1, 7)]
    if actual_headers != expected_headers:
        raise ValidationError(
            f"Unexpected workbook headers: {actual_headers}; expected {expected_headers}"
        )

    audit: list[dict] = []
    updated_rows = 0
    for task in tasks:
        task_id = task["taskId"]
        result = result_by_id[task_id]
        tire_data = result["tireData"]
        metadata = result["researchMetadata"]
        fill = fill_for_score(metadata["reliabilityScore"])
        assets = mapping[task_id]["assets"]

        for asset in assets:
            row = asset["row"]
            if row > worksheet.max_row:
                raise ValidationError(
                    f"{task_id}: mapped row {row} exceeds worksheet row count {worksheet.max_row}"
                )
            worksheet.cell(row=row, column=5).value = tire_data["frontTireSize"]
            worksheet.cell(row=row, column=6).value = tire_data["rearTireSize"]
            worksheet.cell(row=row, column=5).fill = fill
            worksheet.cell(row=row, column=6).fill = fill
            updated_rows += 1

        audit.append(
            {
                "taskId": task_id,
                "vehicle": mapping[task_id].get("vehicle", task["vehicle"]),
                "matchedAssetCount": len(assets),
                "tireData": tire_data,
                "researchMetadata": metadata,
            }
        )

    save_workbook_atomically(workbook, output)
    audit_path.parent.mkdir(parents=True, exist_ok=True)
    fd, temporary_audit = tempfile.mkstemp(
        dir=audit_path.parent, prefix=f".{audit_path.name}.", suffix=".tmp"
    )
    try:
        with os.fdopen(fd, "w", encoding="utf-8") as handle:
            json.dump(audit, handle, indent=2)
            handle.write("\n")
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary_audit, audit_path)
    except Exception:
        try:
            os.unlink(temporary_audit)
        except FileNotFoundError:
            pass
        try:
            output.unlink()
        except FileNotFoundError:
            pass
        raise

    if file_hash(source) != source_hash_before:
        raise RuntimeError("Source workbook changed during injection")
    return len(results), updated_rows


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--source", type=Path, default=DEFAULT_SOURCE)
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT)
    parser.add_argument("--tasks", type=Path, default=DEFAULT_TASKS)
    parser.add_argument("--mapping", type=Path, default=DEFAULT_MAPPING)
    parser.add_argument("--results", type=Path, default=DEFAULT_RESULTS)
    parser.add_argument("--audit", type=Path, default=DEFAULT_AUDIT)
    return parser


def main() -> int:
    args = build_parser().parse_args()
    try:
        task_count, row_count = inject(
            source=args.source,
            output=args.output,
            tasks_path=args.tasks,
            mapping_path=args.mapping,
            results_path=args.results,
            audit_path=args.audit,
        )
    except ValidationError as exc:
        print(f"ERROR: {exc}")
        return 2

    print(f"Created new workbook: {args.output}")
    print(f"Created new audit log: {args.audit}")
    print(f"Research tasks injected: {task_count}")
    print(f"Asset rows updated: {row_count}")
    print(f"Source workbook preserved: {args.source}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
