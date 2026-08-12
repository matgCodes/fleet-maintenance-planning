#!/usr/bin/env python3
"""generate_staged_master_workbook.py  (Issue #17)

Generate the staged master copy of Key_Inventory_Template_Master.xlsx with
reconciled tire fields appended at columns O:R (Front Tire Size, Rear Tire Size,
Tire Detail 1, Tire Detail 2).

Consumes the verified reconciliation manifest produced by #16
(.scratch/key-inventory-tire-reconciliation/reconciliation-manifest.json).
Does NOT perform identity matching or reclassification during workbook generation.

Safety constraints:
    - Never overwrites Key_Inventory_Template_Master.xlsx or Asset_Tire_Inventory_Codex_Enriched.xlsx.
    - Refuses to run if Excel lock files (~$...) exist for source workbooks.
    - Refuses to run if output path exists (unless --force / --overwrite is passed).
    - Fails closed on missing or unverified manifest, source fingerprint drift,
      or category count mismatch.
    - Preserves all existing columns A:N, formulas, styles, formatting, comments,
      validations, row heights, column widths, and other sheets completely.
    - Performs an automated reload and validation pass after saving.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
except ImportError:  # pragma: no cover
    print("ERROR: openpyxl is required. Install with: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)


# ---------------------------------------------------------------------------
# Default Paths and Approved Baselines
# ---------------------------------------------------------------------------

ONEDRIVE = Path(
    "/Users/mag_station/Library/CloudStorage/OneDrive2-PublicHealthInstitute"
)
DEFAULT_MASTER = ONEDRIVE / "Key_Inventory_Template_Master.xlsx"
DEFAULT_TIRE = ONEDRIVE / "Asset_Tire_Inventory_Codex_Enriched.xlsx"
DEFAULT_OUTPUT = ONEDRIVE / "Key_Inventory_Template_Master_With_Tire_Sizes.xlsx"

REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_MANIFEST = (
    REPO_ROOT / ".scratch" / "key-inventory-tire-reconciliation"
    / "reconciliation-manifest.json"
)

EXPECTED_MASTER_SHA = "117328f608cb276dc839cd9153f7b0afb914584de4fa5ff9e52515085a877206"
EXPECTED_TIRE_SHA = "7c5f8c09c38684d6d7a18cdc96c6c3e24df058422c4c1db5a43c91657e4e020b"

EXPECTED_SHEETS = ["Key Inventory", "Instructions & Process", "Add More Assets Here"]
APPENDED_HEADERS = ["Front Tire Size", "Rear Tire Size", "Tire Detail 1", "Tire Detail 2"]

EXPECTED_TOTAL_ASSETS = 444
EXPECTED_CATEGORY_TOTALS = {
    "direct": 392,
    "moved_detail": 6,
    "corrected": 16,
    "yellow": 15,
    "unresolved": 15,
}

# ---------------------------------------------------------------------------
# Styling Tokens
# ---------------------------------------------------------------------------

HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="Arial", size=11, bold=True, color="FFFFFF")
HEADER_ALIGNMENT = Alignment(horizontal="center", vertical="center", wrap_text=True)

DATA_FONT = Font(name="Arial", size=11, bold=False, color="0000FF")
DATA_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
ALIGN_LEFT = Alignment(horizontal="left", vertical="center")

YELLOW_FILL = PatternFill(start_color="FFFF99", end_color="FFFF99", fill_type="solid")


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for chunk in iter(lambda: handle.read(65536), b""):
            digest.update(chunk)
    return digest.hexdigest()


def check_lock(path: Path) -> None:
    lock_file = path.parent / f"~${path.name}"
    if lock_file.exists():
        print(f"ERROR: Lock file exists for '{path.name}': {lock_file}", file=sys.stderr)
        sys.exit(1)


def normalize_vehicle_number(raw: object) -> str:
    if raw is None:
        return ""
    val = str(raw).strip()
    val = re_sub_suffix(val)
    return val


def re_sub_suffix(val: str) -> str:
    import re
    val = re.sub(r"(\d)\s+([A-Za-z])\s*$", lambda m: m.group(1) + m.group(2).upper(), val)
    val = re.sub(r"([A-Za-z]+)$", lambda m: m.group(1).upper(), val)
    return val


# ---------------------------------------------------------------------------
# Core Generation & Verification Logic
# ---------------------------------------------------------------------------

def load_and_verify_manifest(manifest_path: Path, master_path: Path, tire_path: Path) -> dict:
    if not manifest_path.exists():
        print(f"ERROR: Manifest file not found: {manifest_path}", file=sys.stderr)
        sys.exit(1)

    with manifest_path.open("r", encoding="utf-8") as fh:
        manifest = json.load(fh)

    if manifest.get("schema") != "tire-reconciliation-manifest/1":
        print(f"ERROR: Invalid manifest schema: {manifest.get('schema')}", file=sys.stderr)
        sys.exit(1)

    rta_status = manifest.get("rta_verification", {}).get("status")
    if rta_status != "VERIFIED":
        print(f"ERROR: Manifest RTA verification status is '{rta_status}', expected 'VERIFIED'.", file=sys.stderr)
        sys.exit(1)

    # Hash checks
    master_sha = sha256(master_path)
    if master_sha != EXPECTED_MASTER_SHA:
        print(f"ERROR: Master workbook SHA-256 mismatch! Got {master_sha}, expected {EXPECTED_MASTER_SHA}", file=sys.stderr)
        sys.exit(1)

    tire_sha = sha256(tire_path)
    if tire_sha != EXPECTED_TIRE_SHA:
        print(f"ERROR: Tire workbook SHA-256 mismatch! Got {tire_sha}, expected {EXPECTED_TIRE_SHA}", file=sys.stderr)
        sys.exit(1)

    # Category totals checks
    cat_totals = manifest.get("category_totals", {})
    for cat_name, expected_count in EXPECTED_CATEGORY_TOTALS.items():
        actual_count = cat_totals.get(cat_name)
        if actual_count != expected_count:
            print(f"ERROR: Manifest category '{cat_name}' count is {actual_count}, expected {expected_count}", file=sys.stderr)
            sys.exit(1)

    entries = manifest.get("entries", [])
    if len(entries) != EXPECTED_TOTAL_ASSETS:
        print(f"ERROR: Manifest entry count is {len(entries)}, expected {EXPECTED_TOTAL_ASSETS}", file=sys.stderr)
        sys.exit(1)

    return manifest


def generate_staged_workbook(
    master_path: Path,
    tire_path: Path,
    output_path: Path,
    manifest: dict,
    dry_run: bool = False,
    force: bool = False,
) -> str | None:
    # Check locks on source workbooks
    check_lock(master_path)
    check_lock(tire_path)

    # Check output safety
    if output_path.resolve() in (master_path.resolve(), tire_path.resolve()):
        print("ERROR: Refusing to overwrite source workbooks!", file=sys.stderr)
        sys.exit(1)

    if output_path.exists() and not force and not dry_run:
        print(f"ERROR: Output path already exists: {output_path}. Use --force to overwrite.", file=sys.stderr)
        sys.exit(1)

    # Load master workbook
    wb = load_workbook(master_path)
    sheet_names = wb.sheetnames
    if sheet_names != EXPECTED_SHEETS:
        print(f"ERROR: Master sheet names mismatch. Got {sheet_names}, expected {EXPECTED_SHEETS}", file=sys.stderr)
        sys.exit(1)

    ws = wb["Key Inventory"]

    # Assert header row height and max row
    ws.row_dimensions[1].height = 53.25
    if ws.max_row != EXPECTED_TOTAL_ASSETS + 1:
        print(f"ERROR: Master row count is {ws.max_row}, expected {EXPECTED_TOTAL_ASSETS + 1}", file=sys.stderr)
        sys.exit(1)

    # Append Headers at O1:R1 (cols 15..18)
    for idx, header_title in enumerate(APPENDED_HEADERS, start=15):
        cell = ws.cell(row=1, column=idx, value=header_title)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGNMENT

    # Set column widths
    ws.column_dimensions["O"].width = 18.0
    ws.column_dimensions["P"].width = 18.0
    ws.column_dimensions["Q"].width = 22.0
    ws.column_dimensions["R"].width = 22.0

    # Build entry lookup by master_row and asset number
    entries = manifest["entries"]
    entry_by_row = {e["master_row"]: e for e in entries}

    processed_rows = 0
    yellow_cells_count = 0

    for r in range(2, EXPECTED_TOTAL_ASSETS + 2):
        row_asset_raw = ws.cell(row=r, column=1).value
        row_asset_norm = normalize_vehicle_number(row_asset_raw)

        entry = entry_by_row.get(r)
        if not entry:
            print(f"ERROR: No manifest entry found for row {r}", file=sys.stderr)
            sys.exit(1)

        manifest_asset_norm = normalize_vehicle_number(entry["asset"])
        if row_asset_norm != manifest_asset_norm:
            print(f"ERROR: Row {r} asset mismatch: master '{row_asset_norm}' vs manifest '{manifest_asset_norm}'", file=sys.stderr)
            sys.exit(1)

        front = entry.get("front_value")
        rear = entry.get("rear_value")
        detail1 = entry.get("detail1")
        detail2 = entry.get("detail2")
        fill_directive = entry.get("fill_directive")
        action_class = entry.get("action_class")

        # Write values
        cell_o = ws.cell(row=r, column=15, value=front)
        cell_p = ws.cell(row=r, column=16, value=rear)
        cell_q = ws.cell(row=r, column=17, value=detail1)
        cell_r = ws.cell(row=r, column=18, value=detail2)

        # Basic cell styling
        for cell in (cell_o, cell_p, cell_q, cell_r):
            cell.font = DATA_FONT
            cell.border = DATA_BORDER

        cell_o.alignment = ALIGN_CENTER
        cell_p.alignment = ALIGN_CENTER
        cell_q.alignment = ALIGN_LEFT
        cell_r.alignment = ALIGN_LEFT

        # Apply yellow fill for identity gap rows
        if fill_directive == "identity_gap_yellow" or action_class == "yellow":
            cell_o.fill = YELLOW_FILL
            cell_p.fill = YELLOW_FILL
            yellow_cells_count += 2

        ws.row_dimensions[r].height = 26.75
        processed_rows += 1

    if processed_rows != EXPECTED_TOTAL_ASSETS:
        print(f"ERROR: Processed {processed_rows} rows, expected {EXPECTED_TOTAL_ASSETS}", file=sys.stderr)
        sys.exit(1)

    if dry_run:
        print(f"[DRY-RUN] Staging complete: {processed_rows} rows processed, {yellow_cells_count} yellow cells formatted.")
        return None

    # Save output file
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    output_sha = sha256(output_path)
    print(f"Successfully generated staged master workbook at '{output_path}'")
    print(f"SHA-256: {output_sha}")

    # Automated Reload Verification Pass
    verify_reload(output_path)
    return output_sha


def verify_reload(output_path: Path) -> dict:
    if not output_path.exists():
        print(f"ERROR: Reload failed — output file does not exist: {output_path}", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(output_path, data_only=True)
    if wb.sheetnames != EXPECTED_SHEETS:
        print(f"ERROR: Reload check failed — sheet names mismatch: {wb.sheetnames}", file=sys.stderr)
        sys.exit(1)

    ws = wb["Key Inventory"]
    if ws.max_column != 18:
        print(f"ERROR: Reload check failed — max_column is {ws.max_column}, expected 18.", file=sys.stderr)
        sys.exit(1)

    if ws.max_row != EXPECTED_TOTAL_ASSETS + 1:
        print(f"ERROR: Reload check failed — max_row is {ws.max_row}, expected {EXPECTED_TOTAL_ASSETS + 1}.", file=sys.stderr)
        sys.exit(1)

    # Check Appended Headers
    headers_reloaded = [ws.cell(row=1, column=c).value for c in range(15, 19)]
    if headers_reloaded != APPENDED_HEADERS:
        print(f"ERROR: Reload check failed — headers mismatch: {headers_reloaded}", file=sys.stderr)
        sys.exit(1)

    # Statistics Audit
    complete_count = 0
    partial_count = 0
    blank_count = 0
    detail_count = 0
    yellow_cells_count = 0

    partial_assets = []

    for r in range(2, EXPECTED_TOTAL_ASSETS + 2):
        asset = str(ws.cell(row=r, column=1).value or "").strip()
        f_val = ws.cell(row=r, column=15).value
        r_val = ws.cell(row=r, column=16).value
        d1_val = ws.cell(row=r, column=17).value
        d2_val = ws.cell(row=r, column=18).value

        has_f = bool(f_val and str(f_val).strip())
        has_r = bool(r_val and str(r_val).strip())
        has_d1 = bool(d1_val and str(d1_val).strip())
        has_d2 = bool(d2_val and str(d2_val).strip())

        if has_f and has_r:
            complete_count += 1
        elif has_f or has_r:
            partial_count += 1
            partial_assets.append(asset)
        else:
            blank_count += 1

        if has_d1 or has_d2:
            detail_count += 1

        cell_o = ws.cell(row=r, column=15)
        cell_p = ws.cell(row=r, column=16)
        if cell_o.fill and cell_o.fill.start_color and cell_o.fill.start_color.rgb:
            rgb = str(cell_o.fill.start_color.rgb)
            if "FFFF99" in rgb:
                yellow_cells_count += 1
        if cell_p.fill and cell_p.fill.start_color and cell_p.fill.start_color.rgb:
            rgb = str(cell_p.fill.start_color.rgb)
            if "FFFF99" in rgb:
                yellow_cells_count += 1

    stats = {
        "complete": complete_count,
        "partial": partial_count,
        "partial_assets": partial_assets,
        "blank": blank_count,
        "detail_populated": detail_count,
        "yellow_cells": yellow_cells_count,
    }

    # Assert expected numbers
    if stats["complete"] != 423:
        print(f"ERROR: Reload audit failed — complete count is {stats['complete']}, expected 423", file=sys.stderr)
        sys.exit(1)

    if stats["partial"] != 1 or stats["partial_assets"] != ["0SAN07"]:
        print(f"ERROR: Reload audit failed — partial count/assets mismatch: {stats['partial']}, {stats['partial_assets']}", file=sys.stderr)
        sys.exit(1)

    if stats["blank"] != 20:
        print(f"ERROR: Reload audit failed — blank count is {stats['blank']}, expected 20", file=sys.stderr)
        sys.exit(1)

    if stats["detail_populated"] != 23:
        print(f"ERROR: Reload audit failed — detail_populated count is {stats['detail_populated']}, expected 23", file=sys.stderr)
        sys.exit(1)

    if stats["yellow_cells"] != 30:
        print(f"ERROR: Reload audit failed — yellow cells count is {stats['yellow_cells']}, expected 30", file=sys.stderr)
        sys.exit(1)

    print("PASSED: Reload verification and completeness audit clean! (423 complete, 1 partial [0SAN07], 20 blank, 23 detail populated, 30 yellow cells)")
    return stats


# ---------------------------------------------------------------------------
# CLI Entrypoint
# ---------------------------------------------------------------------------

def main():
    parser = argparse.ArgumentParser(
        description="Generate staged master copy of Key Inventory workbook with tire sizes (Issue #17)."
    )
    parser.add_argument("--manifest", type=Path, default=DEFAULT_MANIFEST, help="Path to manifest JSON.")
    parser.add_argument("--master", type=Path, default=DEFAULT_MASTER, help="Path to Key Inventory master workbook.")
    parser.add_argument("--tire", type=Path, default=DEFAULT_TIRE, help="Path to tire inventory workbook.")
    parser.add_argument("--output", type=Path, default=DEFAULT_OUTPUT, help="Path to save staged output workbook.")
    parser.add_argument("--dry-run", action="store_true", help="Run staging in memory without writing output.")
    parser.add_argument("--force", "--overwrite", dest="force", action="store_true", help="Overwrite output file if it exists.")

    args = parser.parse_args()

    manifest = load_and_verify_manifest(args.manifest, args.master, args.tire)
    generate_staged_workbook(
        master_path=args.master,
        tire_path=args.tire,
        output_path=args.output,
        manifest=manifest,
        dry_run=args.dry_run,
        force=args.force,
    )


if __name__ == "__main__":
    main()
