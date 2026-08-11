import openpyxl
from openpyxl.styles import PatternFill
import json
import os

EXCEL_PATH = "/Users/mag_station/Library/CloudStorage/OneDrive2-PublicHealthInstitute/Asset_Tire_Inventory.xlsx"
MAPPING_PATH = "task_asset_mapping.json"
RESULTS_PATH = "research_results.json"
AUDIT_LOG_PATH = "research_audit.json"

# Define Color Fills for Reliability Scores
FILL_HIGH = PatternFill(start_color="D9EAD3", end_color="D9EAD3", fill_type="solid")    # Soft Green (Score 4-5)
FILL_MEDIUM = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # Soft Yellow (Score 3)
FILL_LOW = PatternFill(start_color="FCE5CD", end_color="FCE5CD", fill_type="solid")     # Soft Orange/Red (Score 1-2)

def get_fill_for_score(score):
    try:
        score_val = int(score)
        if score_val >= 4:
            return FILL_HIGH
        elif score_val == 3:
            return FILL_MEDIUM
        else:
            return FILL_LOW
    except (ValueError, TypeError):
        return FILL_LOW

def inject_results(results_file=RESULTS_PATH):
    if not os.path.exists(EXCEL_PATH):
        raise FileNotFoundError(f"Inventory Excel file not found at: {EXCEL_PATH}")
    if not os.path.exists(MAPPING_PATH):
        raise FileNotFoundError(f"Task asset mapping file not found at: {MAPPING_PATH}")
    if not os.path.exists(results_file):
        raise FileNotFoundError(f"Research results file not found at: {results_file}")

    # Load mappings and research results
    with open(MAPPING_PATH, "r", encoding="utf-8") as f:
        task_mapping = json.load(f)

    with open(results_file, "r", encoding="utf-8") as f:
        research_results = json.load(f)

    wb = openpyxl.load_workbook(EXCEL_PATH)
    ws = wb.active

    audit_logs = []
    updated_rows = 0

    for result in research_results:
        task_id = result.get("taskId")
        tire_data = result.get("tireData", {})
        metadata = result.get("researchMetadata", {})

        front_size = tire_data.get("frontTireSize", "")
        rear_size = tire_data.get("rearTireSize", "")
        score = metadata.get("reliabilityScore", 1)

        fill = get_fill_for_score(score)

        if task_id in task_mapping:
            assets = task_mapping[task_id].get("assets", [])
            vehicle_info = task_mapping[task_id].get("vehicle", {})

            for asset in assets:
                row_idx = asset["row"]

                # Write Front Tire (Col E = Col 5) and Rear Tire (Col F = Col 6)
                front_cell = ws.cell(row=row_idx, column=5)
                rear_cell = ws.cell(row=row_idx, column=6)

                front_cell.value = front_size
                rear_cell.value = rear_size

                # Apply reliability color fills
                front_cell.fill = fill
                rear_cell.fill = fill

                updated_rows += 1

            # Log audit entry
            audit_logs.append({
                "taskId": task_id,
                "vehicle": vehicle_info,
                "matchedAssetCount": len(assets),
                "tireData": tire_data,
                "researchMetadata": metadata
            })

    # Save Excel changes
    wb.save(EXCEL_PATH)

    # Save internal audit log
    with open(AUDIT_LOG_PATH, "w", encoding="utf-8") as f:
        json.dump(audit_logs, f, indent=2)

    print(f"✅ Injection & Logging Complete!")
    print(f"  - Tasks Processed: {len(research_results)}")
    print(f"  - Total Excel Rows Updated: {updated_rows}")
    print(f"  - Workbook Saved: {EXCEL_PATH}")
    print(f"  - Audit Log Saved: {AUDIT_LOG_PATH}")

if __name__ == "__main__":
    inject_results()
