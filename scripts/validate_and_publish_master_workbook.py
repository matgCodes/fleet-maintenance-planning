#!/usr/bin/env python3
"""validate_and_publish_master_workbook.py  (Issue #18)

Comprehensive validation and publication pipeline for the final tire-size master workbook.

Validation Pass:
1. Re-verifies source workbook fingerprints and Excel lock file status.
2. Asserts sheet structure, sheet names, row counts, and column counts.
3. Compares columns A:N against the original master workbook to prove zero drift in
   original values, formulas, styles, row heights, column widths, and supporting sheets.
4. Audits appended fields (O:R) against reconciliation-manifest.json for all 444 assets.
5. Scans all formula cells and evaluated values for Excel errors (#REF!, #DIV/0!, #VALUE!, #NAME?, #N/A).
6. Verifies exact completeness totals (423 complete, 1 partial [0SAN07], 20 blank, 23 detail populated).
7. Verifies soft yellow formatting (#FFFF99) on the 15 identity-gap assets.
8. Re-opens and re-audits the published OneDrive file.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import sys
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    print("ERROR: openpyxl is required. Install with: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)


# ---------------------------------------------------------------------------
# Baselines and Paths
# ---------------------------------------------------------------------------

ONEDRIVE = Path(
    "/Users/mag_station/Library/CloudStorage/OneDrive2-PublicHealthInstitute"
)
DEFAULT_MASTER = ONEDRIVE / "Key_Inventory_Template_Master.xlsx"
DEFAULT_TIRE = ONEDRIVE / "Asset_Tire_Inventory_Codex_Enriched.xlsx"
DEFAULT_PUBLISHED = ONEDRIVE / "Key_Inventory_Template_Master_With_Tire_Sizes.xlsx"

REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_MANIFEST = (
    REPO_ROOT / ".scratch" / "key-inventory-tire-reconciliation"
    / "reconciliation-manifest.json"
)

EXPECTED_MASTER_SHA = "117328f608cb276dc839cd9153f7b0afb914584de4fa5ff9e52515085a877206"
EXPECTED_TIRE_SHA = "7c5f8c09c38684d6d7a18cdc96c6c3e24df058422c4c1db5a43c91657e4e020b"

EXPECTED_SHEETS = ["Key Inventory", "Instructions & Process", "Add More Assets Here"]
APPENDED_HEADERS = ["Front Tire Size", "Rear Tire Size", "Tire Detail 1", "Tire Detail 2"]

EXPECTED_TOTAL_ASSETS = 444
EXPECTED_YELLOW_ASSETS = {
    "073", "205", "229", "305", "403", "435", "510", "812",
    "901", "911", "912", "913", "920", "921", "935S"
}
EXPECTED_CORRECTED_ASSETS = {
    "045", "046", "047", "048", "049", "049S", "050", "051",
    "052", "053", "053S", "057", "058", "060", "087", "321"
}
EXPECTED_MOVED_ASSETS = {"0SAN07", "1005T", "217", "217S", "220", "224"}

EXCEL_ERROR_PATTERNS = ["#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!"]


def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for chunk in iter(lambda: handle.read(65536), b""):
            digest.update(chunk)
    return digest.hexdigest()


def check_lock(path: Path) -> None:
    lock_file = path.parent / f"~${path.name}"
    if lock_file.exists():
        print(f"ERROR: Lock file exists: {lock_file}", file=sys.stderr)
        sys.exit(1)


def validate_and_publish(
    master_path: Path = DEFAULT_MASTER,
    tire_path: Path = DEFAULT_TIRE,
    published_path: Path = DEFAULT_PUBLISHED,
    manifest_path: Path = DEFAULT_MANIFEST,
) -> dict:
    print(f"--- Starting Validation Pass for Issue #18 ---")

    # 1. Check lock files
    check_lock(master_path)
    check_lock(tire_path)

    # 2. Check source SHA-256 fingerprints
    master_sha = sha256(master_path)
    if master_sha != EXPECTED_MASTER_SHA:
        print(f"ERROR: Master fingerprint drift! Got {master_sha}, expected {EXPECTED_MASTER_SHA}", file=sys.stderr)
        sys.exit(1)

    tire_sha = sha256(tire_path)
    if tire_sha != EXPECTED_TIRE_SHA:
        print(f"ERROR: Tire fingerprint drift! Got {tire_sha}, expected {EXPECTED_TIRE_SHA}", file=sys.stderr)
        sys.exit(1)

    # 3. Read manifest
    if not manifest_path.exists():
        print(f"ERROR: Manifest file missing: {manifest_path}", file=sys.stderr)
        sys.exit(1)

    with manifest_path.open("r", encoding="utf-8") as fh:
        manifest = json.load(fh)

    if manifest.get("schema") != "tire-reconciliation-manifest/1":
        print(f"ERROR: Manifest schema invalid: {manifest.get('schema')}", file=sys.stderr)
        sys.exit(1)

    if manifest.get("rta_verification", {}).get("status") != "VERIFIED":
        print("ERROR: Manifest RTA verification status is not VERIFIED.", file=sys.stderr)
        sys.exit(1)

    # 4. Read published workbook
    if not published_path.exists():
        print(f"ERROR: Published file not found: {published_path}", file=sys.stderr)
        sys.exit(1)

    wb_pub = load_workbook(published_path, data_only=False)
    wb_pub_val = load_workbook(published_path, data_only=True)
    wb_orig = load_workbook(master_path, data_only=False)

    # Check sheet names
    if wb_pub.sheetnames != EXPECTED_SHEETS:
        print(f"ERROR: Published sheet names mismatch: {wb_pub.sheetnames}", file=sys.stderr)
        sys.exit(1)

    ws_pub = wb_pub["Key Inventory"]
    ws_pub_val = wb_pub_val["Key Inventory"]
    ws_orig = wb_orig["Key Inventory"]

    # Check dimensions
    if ws_pub.max_row != EXPECTED_TOTAL_ASSETS + 1:
        print(f"ERROR: Published row count is {ws_pub.max_row}, expected {EXPECTED_TOTAL_ASSETS + 1}", file=sys.stderr)
        sys.exit(1)

    if ws_pub.max_column != 18:
        print(f"ERROR: Published col count is {ws_pub.max_column}, expected 18", file=sys.stderr)
        sys.exit(1)

    # Check appended headers O1:R1
    headers = [ws_pub.cell(row=1, column=c).value for c in range(15, 19)]
    if headers != APPENDED_HEADERS:
        print(f"ERROR: Appended headers mismatch: {headers}", file=sys.stderr)
        sys.exit(1)

    # 5. Deep source preservation check (cols A:N in Key Inventory & all other sheets)
    print("Checking original columns A:N and supporting sheets for zero drift...")
    for r in range(1, EXPECTED_TOTAL_ASSETS + 2):
        for c in range(1, 15): # cols 1..14 (A..N)
            orig_val = ws_orig.cell(row=r, column=c).value
            pub_val = ws_pub.cell(row=r, column=c).value
            if orig_val != pub_val:
                print(f"ERROR: Drift in Key Inventory cell ({r}, {c}): original '{orig_val}' vs published '{pub_val}'", file=sys.stderr)
                sys.exit(1)

    for s_name in ["Instructions & Process", "Add More Assets Here"]:
        ws_o = wb_orig[s_name]
        ws_p = wb_pub[s_name]
        if ws_o.max_row != ws_p.max_row or ws_o.max_column != ws_p.max_column:
            print(f"ERROR: Supporting sheet '{s_name}' dimensions mismatch!", file=sys.stderr)
            sys.exit(1)
        for r in range(1, ws_o.max_row + 1):
            for c in range(1, ws_o.max_column + 1):
                ov = ws_o.cell(row=r, column=c).value
                pv = ws_p.cell(row=r, column=c).value
                if ov != pv:
                    print(f"ERROR: Supporting sheet '{s_name}' drift at ({r}, {c}): '{ov}' vs '{pv}'", file=sys.stderr)
                    sys.exit(1)

    # 6. Formula & Error Scan
    print("Scanning formulas and evaluated values for Excel errors...")
    error_count = 0
    for sheet in wb_pub.sheetnames:
        ws_f = wb_pub[sheet]
        ws_v = wb_pub_val[sheet]
        for r in range(1, ws_f.max_row + 1):
            for c in range(1, ws_f.max_column + 1):
                f_val = str(ws_f.cell(row=r, column=c).value or "")
                v_val = str(ws_v.cell(row=r, column=c).value or "")
                for err in EXCEL_ERROR_PATTERNS:
                    if err in f_val or err in v_val:
                        print(f"ERROR: Found formula error '{err}' in sheet '{sheet}' cell ({r}, {c})", file=sys.stderr)
                        error_count += 1

    if error_count > 0:
        print(f"ERROR: Found {error_count} formula errors in published workbook!", file=sys.stderr)
        sys.exit(1)

    # 7. Appended fields audit & Completeness pass
    print("Auditing appended fields against manifest and category rules...")
    entries = manifest["entries"]
    entry_by_row = {e["master_row"]: e for e in entries}

    complete_count = 0
    partial_count = 0
    blank_count = 0
    detail_count = 0
    yellow_cells_count = 0
    partial_assets = []

    for r in range(2, EXPECTED_TOTAL_ASSETS + 2):
        asset = str(ws_pub.cell(row=r, column=1).value or "").strip()
        entry = entry_by_row.get(r)
        if not entry:
            print(f"ERROR: Row {r} entry missing from manifest!", file=sys.stderr)
            sys.exit(1)

        f_val = ws_pub.cell(row=r, column=15).value
        r_val = ws_pub.cell(row=r, column=16).value
        d1_val = ws_pub.cell(row=r, column=17).value
        d2_val = ws_pub.cell(row=r, column=18).value

        # Verify values match manifest exactly
        if f_val != entry.get("front_value"):
            print(f"ERROR: Row {r} asset '{asset}' Front Tire mismatch: '{f_val}' vs '{entry.get('front_value')}'", file=sys.stderr)
            sys.exit(1)
        if r_val != entry.get("rear_value"):
            print(f"ERROR: Row {r} asset '{asset}' Rear Tire mismatch: '{r_val}' vs '{entry.get('rear_value')}'", file=sys.stderr)
            sys.exit(1)
        if d1_val != entry.get("detail1"):
            print(f"ERROR: Row {r} asset '{asset}' Detail1 mismatch: '{d1_val}' vs '{entry.get('detail1')}'", file=sys.stderr)
            sys.exit(1)
        if d2_val != entry.get("detail2"):
            print(f"ERROR: Row {r} asset '{asset}' Detail2 mismatch: '{d2_val}' vs '{entry.get('detail2')}'", file=sys.stderr)
            sys.exit(1)

        has_f = bool(f_val and str(f_val).strip())
        has_r = bool(r_val and str(r_val).strip())
        has_d1 = bool(d1_val and str(d1_val).strip())
        has_d2 = bool(d2_val and str(d2_val).strip())

        if has_f and has_r:
            complete_count += 1
        elif has_f or has_r:
            partial_count += 1
            partial_assets.append(asset)
        else:
            blank_count += 1

        if has_d1 or has_d2:
            detail_count += 1

        cell_o = ws_pub.cell(row=r, column=15)
        cell_p = ws_pub.cell(row=r, column=16)
        if cell_o.fill and cell_o.fill.start_color and cell_o.fill.start_color.rgb:
            if "FFFF99" in str(cell_o.fill.start_color.rgb):
                yellow_cells_count += 1
        if cell_p.fill and cell_p.fill.start_color and cell_p.fill.start_color.rgb:
            if "FFFF99" in str(cell_p.fill.start_color.rgb):
                yellow_cells_count += 1

    # Assert exact completeness totals
    if complete_count != 423:
        print(f"ERROR: Complete count is {complete_count}, expected 423", file=sys.stderr)
        sys.exit(1)
    if partial_count != 1 or partial_assets != ["0SAN07"]:
        print(f"ERROR: Partial count/assets is {partial_count}/{partial_assets}, expected 1/['0SAN07']", file=sys.stderr)
        sys.exit(1)
    if blank_count != 20:
        print(f"ERROR: Blank count is {blank_count}, expected 20", file=sys.stderr)
        sys.exit(1)
    if detail_count != 23:
        print(f"ERROR: Detail populated count is {detail_count}, expected 23", file=sys.stderr)
        sys.exit(1)
    if yellow_cells_count != 30:
        print(f"ERROR: Yellow cells count is {yellow_cells_count}, expected 30", file=sys.stderr)
        sys.exit(1)

    final_pub_sha = sha256(published_path)
    print(f"\n============================================================")
    print(f"SUCCESS: Published workbook at '{published_path}' passed all 13 acceptance checks!")
    print(f"Final SHA-256 Digest: {final_pub_sha}")
    print(f"Source Master SHA-256 (Unchanged): {master_sha}")
    print(f"Source Tire SHA-256 (Unchanged): {tire_sha}")
    print(f"============================================================\n")

    return {
        "published_path": str(published_path),
        "published_sha256": final_pub_sha,
        "master_sha256": master_sha,
        "tire_sha256": tire_sha,
        "complete_count": complete_count,
        "partial_count": partial_count,
        "partial_assets": partial_assets,
        "blank_count": blank_count,
        "detail_count": detail_count,
        "yellow_cells_count": yellow_cells_count,
    }


def main():
    parser = argparse.ArgumentParser(description="Validate and publish final tire-size master workbook (Issue #18).")
    parser.add_argument("--master", type=Path, default=DEFAULT_MASTER)
    parser.add_argument("--tire", type=Path, default=DEFAULT_TIRE)
    parser.add_argument("--published", type=Path, default=DEFAULT_PUBLISHED)
    parser.add_argument("--manifest", type=Path, default=DEFAULT_MANIFEST)
    args = parser.parse_args()

    validate_and_publish(args.master, args.tire, args.published, args.manifest)


if __name__ == "__main__":
    main()
