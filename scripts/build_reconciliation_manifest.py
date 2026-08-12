#!/usr/bin/env python3
"""build_reconciliation_manifest.py  (Issue #16)

Deterministic, read-only tire-reconciliation manifest builder.

Joins every Key Inventory asset to the authoritative tire workbook, re-verifies
the master's identity against live RTA aggregates (read from stdin, never
persisted), classifies every asset into exactly one approved action set, and
emits a frozen manifest that the workbook-generation ticket (#17) can consume
without making any further identity or precedence decisions.

The Key Inventory master stays authoritative for Asset #, Year, Make, Model, and
VIN. The tire workbook and Codex research populate only the four appended tire
fields and their presentation.

Usage:
    ./scripts/rta-reconcile-fetch.mjs | python3 ./scripts/build_reconciliation_manifest.py
  OR (offline / tests):
    python3 ./scripts/build_reconciliation_manifest.py --rta-json /path/to/vehicles.json
  OR (local classification only, no RTA):
    python3 ./scripts/build_reconciliation_manifest.py --skip-rta

Flags:
    --dry-run        Run every gate + classification but do not write the manifest.
    --rta-json PATH  Read RTA vehicle JSON from a file instead of stdin.
    --skip-rta       Skip the live RTA re-verification (manifest is tagged SKIPPED
                     and must NOT be treated as complete for #17). For local checks.
    --master PATH    Override the master workbook path (tests).
    --tire PATH      Override the tire workbook path (tests).
    --output PATH    Override the final publish path checked for non-existence (tests).
    --manifest PATH  Override the manifest output path (tests).
    --map-json PATH  Override task_asset_mapping.json (tests).
    --codex-json PATH Override codex_research_results.json (tests).
    --help           Show this message.

Safety constraints (inherited from AGENTS.md + issues #16-#18):
    - Read-only: never modifies either source workbook or any repo copy.
    - Fails closed (non-zero exit) on any lock file, source-hash drift, existing
      output path, join anomaly, category-membership/count drift, or RTA-aggregate
      drift, before writing anything.
    - RTA vehicle data (incl. serialNumber/VIN) is accepted only via stdin/file,
      is never persisted or printed. Only aggregate counts and business
      asset-number exception lists are printed or written.
    - The emitted manifest contains asset numbers, tire sizes, and aggregate
      counts only -- never VINs, serials, tokens, or raw RTA records.
"""

from __future__ import annotations

import argparse
import datetime as dt
import hashlib
import json
import os
import re
import sys
from collections import defaultdict
from pathlib import Path

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    print("ERROR: openpyxl is required. Install with: pip3 install openpyxl", file=sys.stderr)
    sys.exit(1)


# ---------------------------------------------------------------------------
# Paths (authoritative sources live in OneDrive, not the repo root)
# ---------------------------------------------------------------------------

ONEDRIVE = Path(
    "/Users/mag_station/Library/CloudStorage/OneDrive2-PublicHealthInstitute"
)
DEFAULT_MASTER = ONEDRIVE / "Key_Inventory_Template_Master.xlsx"
DEFAULT_TIRE = ONEDRIVE / "Asset_Tire_Inventory_Codex_Enriched.xlsx"
DEFAULT_OUTPUT = ONEDRIVE / "Key_Inventory_Template_Master_With_Tire_Sizes.xlsx"

REPO_ROOT = Path(__file__).resolve().parent.parent
DEFAULT_MAP_JSON = REPO_ROOT / "task_asset_mapping.json"
DEFAULT_CODEX_JSON = REPO_ROOT / "codex_research_results.json"
DEFAULT_MANIFEST = (
    REPO_ROOT / ".scratch" / "key-inventory-tire-reconciliation"
    / "reconciliation-manifest.json"
)

# ---------------------------------------------------------------------------
# Approved baseline (frozen; the fail-closed contract for #16)
# ---------------------------------------------------------------------------

# Source fingerprints of the approved authoritative sources.
EXPECTED_MASTER_SHA = "117328f608cb276dc839cd9153f7b0afb914584de4fa5ff9e52515085a877206"
EXPECTED_TIRE_SHA = "7c5f8c09c38684d6d7a18cdc96c6c3e24df058422c4c1db5a43c91657e4e020b"

EXPECTED_SHEETS = ["Key Inventory", "Instructions & Process", "Add More Assets Here"]

# Current master layout (14 cols): VIN was inserted at D, a blank header sits at
# L, and the Total Found / Shortage formulas moved to I and K. Columns are
# resolved from the header row at runtime; these names/positions are asserted.
EXPECTED_MASTER_HEADERS = [
    "Asset #", "Year", "Make", "VIN", "Model", "Master Box", "Fleet Manager",
    "Tech Hooks", "Total Found", "Target Copies", "Shortage", None, "Status", "Notes",
]
EXPECTED_TIRE_HEADERS_A_F = [
    "Asset ID Number", "Year", "Make", "Model", "Front Tire Size", "Rear Tire Size",
]
# Appended tire fields land after Notes at O:R (cols 15-18) in #17.
APPENDED_FIELDS = ["Front Tire Size", "Rear Tire Size", "Tire Detail 1", "Tire Detail 2"]
APPENDED_START_COL = 15  # O

EXPECTED_ASSET_ROWS = 444

# --- RTA aggregate baseline (AC#3/AC#4) ---
APPROVED_AMBIGUOUS = {"014", "021", "169"}
APPROVED_UNMATCHED = {
    "156S", "173", "180", "231", "261", "305", "314", "318", "321S", "329",
    "336", "510", "705", "771", "772", "774", "802", "812", "901", "906",
    "911", "912", "913", "920", "921", "932", "935S", "942",
}
RTA_MATCHED = 413
RTA_YMM_MATCH = 412
RTA_YMM_MISMATCH = 1
RTA_VIN_EXACT = 411
RTA_VIN_CONFLICT = 0
# The single approved residual YMM mismatch, pinned from the live re-verification
# (asset 073 is also an approved yellow identity-gap row). Populated -> the gate
# checks membership as well as count.
APPROVED_YMM_MISMATCH: set[str] = {"073"}

# --- Category action sets (AC#5/AC#6); memberships verified to reproduce the
#     approved totals from the current sources. Derived independently at runtime
#     and asserted equal to these frozen sets. ---
APPROVED_MOVED_DETAIL = {"0SAN07", "1005T", "217", "217S", "220", "224"}
APPROVED_UNRESOLVED = {
    "173", "180", "231", "261", "314", "329", "336", "705",
    "771", "772", "774", "802", "906", "932", "942",
}
APPROVED_CORRECTED = {
    "045", "046", "047", "048", "049", "049S", "050", "051",
    "052", "053", "053S", "057", "058", "060", "087", "321",
}
APPROVED_YELLOW = {
    "073", "205", "229", "305", "403", "435", "510", "812",
    "901", "911", "912", "913", "920", "921", "935S",
}
TOTAL_DIRECT = 392
TOTAL_MOVED = 6
TOTAL_CORRECTED = 16
TOTAL_YELLOW = 15
TOTAL_UNRESOLVED = 15

# --- Tire completeness of the final assignment (cross-checks #18) ---
EXPECTED_COMPLETE = 423
EXPECTED_PARTIAL = 1
EXPECTED_BLANK = 20
EXPECTED_DETAIL_POPULATED = 23  # 6 moved + 16 corrected + 1 (asset 229 yellow)

FILL_YELLOW = "identity_gap_yellow"


# ---------------------------------------------------------------------------
# Small helpers
# ---------------------------------------------------------------------------

def sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with Path(path).open("rb") as handle:
        for chunk in iter(lambda: handle.read(65536), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalize_vehicle_number(raw) -> str:
    """Trim, collapse whitespace before a trailing suffix letter, uppercase the
    suffix, preserve leading zeros. '158 s' -> '158S'; '318' != '318S'."""
    if raw is None:
        return ""
    value = str(raw).strip()
    value = re.sub(r"(\d)\s+([A-Za-z])\s*$",
                   lambda m: m.group(1) + m.group(2).upper(), value)
    value = re.sub(r"([A-Za-z]+)$", lambda m: m.group(1).upper(), value)
    return value


def lookup_keys(norm: str) -> set[str]:
    keys = {norm}
    if norm.isdigit():
        stripped = norm.lstrip("0") or "0"
        keys.update({stripped, norm.zfill(3), stripped.zfill(3)})
    return keys


def clean(value) -> str:
    return "" if value is None else str(value).strip()


def norm_ymm(year, make, model) -> tuple[str, str, str]:
    def n(x: object) -> str:
        s = clean(x).upper()
        return re.sub(r"\s+", " ", s)
    y = n(year)
    m = re.match(r"^(\d{4})(\.0)?$", y)
    if m:
        y = m.group(1)
    return (y, n(make), n(model))


def fail(label: str, errors: list[str]) -> "NoReturn":  # type: ignore[name-defined]
    print(f"\nERROR: {label}", file=sys.stderr)
    for err in errors:
        print(f"  - {err}", file=sys.stderr)
    raise SystemExit(1)


def diff_report(label: str, actual: set[str], expected: set[str]) -> str:
    missing = sorted(expected - actual)
    extra = sorted(actual - expected)
    return f"{label} membership drift: missing={missing} extra={extra}"


# ---------------------------------------------------------------------------
# Column resolution (derive from header row; never hardcode indices)
# ---------------------------------------------------------------------------

def resolve_master_columns(ws) -> dict:
    if ws.max_row != 445 or ws.max_column != 14:
        fail("master dimensions", [f"{ws.max_row}x{ws.max_column}, expected 445x14"])
    headers = [ws.cell(1, c).value for c in range(1, 15)]
    if headers != EXPECTED_MASTER_HEADERS:
        fail("master headers", [f"{headers} != {EXPECTED_MASTER_HEADERS}"])
    if headers[11] is not None:
        fail("master blank column", [f"expected blank header at column 12 (L), got {headers[11]!r}"])
    cols = {name: idx + 1 for idx, name in enumerate(headers) if name is not None}
    # Formulas: Total Found (I) and Shortage (K), 443 each.
    for name in ("Total Found", "Shortage"):
        col = cols[name]
        n = sum(
            1 for r in range(2, 446)
            if isinstance(ws.cell(r, col).value, str) and ws.cell(r, col).value.startswith("=")
        )
        if n != 443:
            fail("master formula count", [f"{name} (col {col}) has {n} formulas, expected 443"])
    return cols


def resolve_tire_columns(ws) -> dict:
    headers = [ws.cell(1, c).value for c in range(1, 7)]
    if headers != EXPECTED_TIRE_HEADERS_A_F:
        fail("tire headers", [f"{headers} != {EXPECTED_TIRE_HEADERS_A_F}"])
    return {
        "asset": 1, "year": 2, "make": 3, "model": 4,
        "front": 5, "rear": 6, "detail1": 7, "detail2": 8,
    }


# ---------------------------------------------------------------------------
# Source readers
# ---------------------------------------------------------------------------

def read_master(ws, cols: dict) -> dict:
    """asset_norm -> {row, ymm, vin}. VIN is held in memory only for the RTA
    re-verification; it is never printed or written to the manifest."""
    out: dict[str, dict] = {}
    for r in range(2, 446):
        asset_raw = ws.cell(r, cols["Asset #"]).value
        if asset_raw is None or clean(asset_raw) == "":
            continue
        norm = normalize_vehicle_number(asset_raw)
        if norm in out:
            fail("master duplicate asset", [f"asset {norm} appears more than once"])
        out[norm] = {
            "row": r,
            "ymm": norm_ymm(ws.cell(r, cols["Year"]).value,
                            ws.cell(r, cols["Make"]).value,
                            ws.cell(r, cols["Model"]).value),
            "vin": clean(ws.cell(r, cols["VIN"]).value),
        }
    return out


def read_tire(ws, cols: dict) -> dict:
    out: dict[str, dict] = {}
    for r in range(2, ws.max_row + 1):
        asset_raw = ws.cell(r, cols["asset"]).value
        if asset_raw is None or clean(asset_raw) == "":
            continue
        norm = normalize_vehicle_number(asset_raw)
        if norm in out:
            fail("tire duplicate asset", [f"asset {norm} appears more than once"])
        out[norm] = {
            "ymm": norm_ymm(ws.cell(r, cols["year"]).value,
                            ws.cell(r, cols["make"]).value,
                            ws.cell(r, cols["model"]).value),
            "front": clean(ws.cell(r, cols["front"]).value),
            "rear": clean(ws.cell(r, cols["rear"]).value),
            "detail1": clean(ws.cell(r, cols["detail1"]).value),
            "detail2": clean(ws.cell(r, cols["detail2"]).value),
        }
    return out


def build_codex_ymm_map(map_json: Path, codex_json: Path) -> tuple[dict, set[str]]:
    """Return (ymm -> {front, rear}) for YMM groups with a usable Codex result,
    and the set of normalized asset numbers that appear in task_asset_mapping."""
    tam = json.loads(Path(map_json).read_text())
    codex = json.loads(Path(codex_json).read_text())
    res_by_task = {rec.get("taskId"): rec for rec in codex if rec.get("taskId")}

    ymm_map: dict[tuple, dict] = {}
    mapped_assets: set[str] = set()
    for task_id, grp in tam.items():
        for a in grp.get("assets", []):
            mapped_assets.add(normalize_vehicle_number(a.get("assetId")))
        veh = grp.get("vehicle", {})
        ymm = norm_ymm(veh.get("year"), veh.get("make"), veh.get("model"))
        rec = res_by_task.get(task_id)
        td = (rec or {}).get("tireData") or {}
        front = clean(td.get("frontTireSize"))
        rear = clean(td.get("rearTireSize"))
        if front:
            ymm_map[ymm] = {"front": front, "rear": rear}
    return ymm_map, mapped_assets


# ---------------------------------------------------------------------------
# Join + classification (fully local; no RTA required)
# ---------------------------------------------------------------------------

def join_and_classify(master: dict, tire: dict, ymm_map: dict, mapped_assets: set[str]) -> dict:
    # --- Join (AC#2): every master asset to exactly one tire row ---
    master_assets = set(master)
    tire_assets = set(tire)
    if len(master) != EXPECTED_ASSET_ROWS:
        fail("master row count", [f"{len(master)} asset rows, expected {EXPECTED_ASSET_ROWS}"])
    only_master = sorted(master_assets - tire_assets)
    only_tire = sorted(tire_assets - master_assets)
    if only_master or only_tire:
        fail("join anomaly", [f"master-only={only_master}", f"tire-only={only_tire}"])

    # --- Unresolved (AC#5): master assets absent from task_asset_mapping ---
    unresolved = {a for a in master_assets if a not in mapped_assets}

    # --- Disagreement pool: master YMM != tire YMM ---
    pool = {a for a in master_assets if master[a]["ymm"] != tire[a]["ymm"]}
    pool -= unresolved  # unresolved rows carry no researched identity

    corrected = {a for a in pool if master[a]["ymm"] in ymm_map}
    yellow = pool - corrected

    moved = set(APPROVED_MOVED_DETAIL)
    direct = master_assets - unresolved - moved - corrected - yellow

    classes = {
        "direct": direct, "moved_detail": moved, "corrected": corrected,
        "yellow": yellow, "unresolved": unresolved,
    }

    # --- Exclusivity + coverage ---
    errors: list[str] = []
    names = list(classes)
    for i in range(len(names)):
        for j in range(i + 1, len(names)):
            overlap = classes[names[i]] & classes[names[j]]
            if overlap:
                errors.append(f"overlap {names[i]}&{names[j]}={sorted(overlap)}")
    union = set().union(*classes.values())
    if union != master_assets:
        errors.append(diff_report("category coverage", union, master_assets))
    if errors:
        fail("category exclusivity/coverage", errors)

    # --- Exact totals (AC#5) ---
    total_errors = []
    for name, count in (("direct", TOTAL_DIRECT), ("moved_detail", TOTAL_MOVED),
                        ("corrected", TOTAL_CORRECTED), ("yellow", TOTAL_YELLOW),
                        ("unresolved", TOTAL_UNRESOLVED)):
        if len(classes[name]) != count:
            total_errors.append(f"{name}={len(classes[name])}, expected {count}")
    if total_errors:
        fail("category totals", total_errors)

    # --- Frozen-membership gate (AC#6: fail closed on membership drift) ---
    member_errors = []
    for name, approved in (("moved_detail", APPROVED_MOVED_DETAIL),
                           ("corrected", APPROVED_CORRECTED),
                           ("yellow", APPROVED_YELLOW),
                           ("unresolved", APPROVED_UNRESOLVED)):
        if classes[name] != approved:
            member_errors.append(diff_report(name, classes[name], approved))
    if member_errors:
        fail("category membership", member_errors)

    return classes


# ---------------------------------------------------------------------------
# RTA re-verification (AC#3/AC#4); RTA read from stdin/file, aggregate-only out
# ---------------------------------------------------------------------------

def verify_rta(master: dict, vehicles: list[dict]) -> dict:
    ids = [v.get("id") for v in vehicles]
    if any(i is None for i in ids) or len(ids) != len(set(ids)):
        fail("RTA integrity", ["RTA ID uniqueness check failed"])

    rta_map: dict[str, list[dict]] = defaultdict(list)
    for v in vehicles:
        for key in lookup_keys(normalize_vehicle_number(v.get("vehicleNumber"))):
            rta_map[key].append(v)

    unmatched, ambiguous = set(), set()
    ymm_match, ymm_mismatch = set(), set()
    vin_exact = vin_conflict = 0

    for asset, info in master.items():
        by_id = {}
        for key in lookup_keys(asset):
            for v in rta_map.get(key, []):
                by_id[v.get("id")] = v
        matches = list(by_id.values())
        if not matches:
            unmatched.add(asset)
            continue
        if len(matches) > 1:
            ambiguous.add(asset)
            continue
        v = matches[0]
        if norm_ymm(v.get("year"), v.get("make"), v.get("model")) == info["ymm"]:
            ymm_match.add(asset)
        else:
            ymm_mismatch.add(asset)
        rta_serial = clean(v.get("serialNumber")).upper()
        master_vin = info["vin"].upper()
        if master_vin and rta_serial:
            if master_vin == rta_serial:
                vin_exact += 1
            else:
                vin_conflict += 1

    matched = len(master) - len(unmatched) - len(ambiguous)
    errors: list[str] = []
    if ambiguous != APPROVED_AMBIGUOUS:
        errors.append(diff_report("ambiguous", ambiguous, APPROVED_AMBIGUOUS))
    if unmatched != APPROVED_UNMATCHED:
        errors.append(diff_report("unmatched", unmatched, APPROVED_UNMATCHED))
    if matched != RTA_MATCHED:
        errors.append(f"matched={matched}, expected {RTA_MATCHED}")
    if len(ymm_match) != RTA_YMM_MATCH:
        errors.append(f"ymm_match={len(ymm_match)}, expected {RTA_YMM_MATCH}")
    if len(ymm_mismatch) != RTA_YMM_MISMATCH:
        errors.append(f"ymm_mismatch={len(ymm_mismatch)}, expected {RTA_YMM_MISMATCH}")
    if APPROVED_YMM_MISMATCH and ymm_mismatch != APPROVED_YMM_MISMATCH:
        errors.append(diff_report("ymm_mismatch", ymm_mismatch, APPROVED_YMM_MISMATCH))
    if vin_exact != RTA_VIN_EXACT:
        errors.append(f"vin_exact={vin_exact}, expected {RTA_VIN_EXACT}")
    if vin_conflict != RTA_VIN_CONFLICT:
        errors.append(f"vin_conflict={vin_conflict}, expected {RTA_VIN_CONFLICT}")
    if errors:
        fail("RTA aggregate baseline", errors)

    return {
        "status": "VERIFIED",
        "rta_vehicles": len(vehicles),
        "matched": matched,
        "ambiguous": len(ambiguous),
        "unmatched": len(unmatched),
        "ymm_match": len(ymm_match),
        "ymm_mismatch": len(ymm_mismatch),
        "ymm_mismatch_assets": sorted(ymm_mismatch),
        "vin_exact": vin_exact,
        "vin_conflict": vin_conflict,
    }


# ---------------------------------------------------------------------------
# Manifest assembly
# ---------------------------------------------------------------------------

def build_entries(master: dict, tire: dict, ymm_map: dict, classes: dict) -> list[dict]:
    asset_class = {a: name for name, members in classes.items() for a in members}
    entries = []
    for asset, info in master.items():
        cls = asset_class[asset]
        t = tire[asset]
        front = rear = detail1 = detail2 = None
        fill = None
        if cls == "direct":
            front, rear = t["front"] or None, t["rear"] or None
        elif cls == "moved_detail":
            front, rear = t["front"] or None, t["rear"] or None
            detail1, detail2 = t["detail1"] or None, t["detail2"] or None
        elif cls == "corrected":
            res = ymm_map[master[asset]["ymm"]]
            front, rear = res["front"] or None, res["rear"] or None
            detail1, detail2 = t["front"] or None, t["rear"] or None  # superseded current
        elif cls == "yellow":
            front, rear = t["front"] or None, t["rear"] or None
            detail1, detail2 = t["detail1"] or None, t["detail2"] or None
            fill = FILL_YELLOW
        elif cls == "unresolved":
            pass  # all four blank
        entries.append({
            "master_row": info["row"],
            "asset": asset,
            "action_class": cls,
            "front_value": front,
            "rear_value": rear,
            "detail1": detail1,
            "detail2": detail2,
            "fill_directive": fill,
        })
    entries.sort(key=lambda e: e["master_row"])
    return entries


def check_completeness(entries: list[dict]) -> dict:
    complete = partial = blank = detail_pop = 0
    partial_assets = []
    for e in entries:
        f = bool(e["front_value"])
        r = bool(e["rear_value"])
        if f and r:
            complete += 1
        elif not f and not r:
            blank += 1
        else:
            partial += 1
            partial_assets.append(e["asset"])
        if e["detail1"] or e["detail2"]:
            detail_pop += 1
    errors = []
    if complete != EXPECTED_COMPLETE:
        errors.append(f"complete={complete}, expected {EXPECTED_COMPLETE}")
    if partial != EXPECTED_PARTIAL:
        errors.append(f"partial={partial}, expected {EXPECTED_PARTIAL}")
    if blank != EXPECTED_BLANK:
        errors.append(f"blank={blank}, expected {EXPECTED_BLANK}")
    if detail_pop != EXPECTED_DETAIL_POPULATED:
        errors.append(f"detail_populated={detail_pop}, expected {EXPECTED_DETAIL_POPULATED}")
    if errors:
        fail("tire completeness cross-check", errors)
    return {"complete": complete, "partial": partial, "blank": blank,
            "detail_populated": detail_pop, "partial_assets": partial_assets}


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def parse_args() -> argparse.Namespace:
    p = argparse.ArgumentParser(add_help=False)
    p.add_argument("--dry-run", action="store_true")
    p.add_argument("--skip-rta", action="store_true")
    p.add_argument("--rta-json")
    p.add_argument("--master", default=str(DEFAULT_MASTER))
    p.add_argument("--tire", default=str(DEFAULT_TIRE))
    p.add_argument("--output", default=str(DEFAULT_OUTPUT))
    p.add_argument("--manifest", default=str(DEFAULT_MANIFEST))
    p.add_argument("--map-json", default=str(DEFAULT_MAP_JSON))
    p.add_argument("--codex-json", default=str(DEFAULT_CODEX_JSON))
    p.add_argument("--help", "-h", action="store_true")
    return p.parse_args()


def main() -> None:
    args = parse_args()
    if args.help:
        print(__doc__)
        return

    master_path = Path(args.master)
    tire_path = Path(args.tire)
    output_path = Path(args.output)
    manifest_path = Path(args.manifest)
    mode = "dry-run (no manifest write)" if args.dry_run else "build"
    print(f"MODE: {mode}")

    # --- 1. Source gates (AC#1) ---
    for label, p in (("master", master_path), ("tire", tire_path)):
        lock = p.with_name(f"~${p.name}")
        if lock.exists():
            fail("lock file", [f"{label} lock present: {lock}"])
    print("lock_files=clear")

    master_sha = sha256(master_path)
    tire_sha = sha256(tire_path)
    print(f"master_sha256={master_sha}")
    print(f"tire_sha256={tire_sha}")
    hash_errors = []
    if master_sha != EXPECTED_MASTER_SHA:
        hash_errors.append(f"master expected {EXPECTED_MASTER_SHA}, got {master_sha}")
    if tire_sha != EXPECTED_TIRE_SHA:
        hash_errors.append(f"tire expected {EXPECTED_TIRE_SHA}, got {tire_sha}")
    if hash_errors:
        fail("source fingerprint", hash_errors)
    print("source_fingerprints=match")

    if output_path.exists():
        fail("output path", [f"final output already exists: {output_path}"])
    print("output_path=absent")

    # --- 2. Load sources (read-only) ---
    mwb = load_workbook(master_path, data_only=False)
    if mwb.sheetnames != EXPECTED_SHEETS:
        fail("master sheets", [f"{mwb.sheetnames} != {EXPECTED_SHEETS}"])
    mws = mwb["Key Inventory"]
    mcols = resolve_master_columns(mws)
    master = read_master(mws, mcols)
    print(f"master_rows={len(master)}")

    twb = load_workbook(tire_path, data_only=False)
    tws = twb.active
    tcols = resolve_tire_columns(tws)
    tire = read_tire(tws, tcols)
    print(f"tire_rows={len(tire)}")

    ymm_map, mapped_assets = build_codex_ymm_map(Path(args.map_json), Path(args.codex_json))
    print(f"codex_ymm_groups={len(ymm_map)} mapped_assets={len(mapped_assets)}")

    # --- 3. Join + classify (AC#2/AC#5/AC#6) ---
    classes = join_and_classify(master, tire, ymm_map, mapped_assets)
    print("join=444/444 category_gate=PASS")
    print(f"  direct={len(classes['direct'])} moved_detail={len(classes['moved_detail'])} "
          f"corrected={len(classes['corrected'])} yellow={len(classes['yellow'])} "
          f"unresolved={len(classes['unresolved'])}")

    # --- 4. RTA re-verification (AC#3/AC#4) ---
    if args.skip_rta:
        rta = {"status": "SKIPPED"}
        print("rta_verification=SKIPPED (manifest NOT complete for #17)")
    else:
        if args.rta_json:
            vehicles = json.loads(Path(args.rta_json).read_text())
        else:
            raw = sys.stdin.read()
            if not raw.strip():
                fail("RTA input", ["no RTA JSON on stdin; use --rta-json or --skip-rta"])
            vehicles = json.loads(raw)
        if not isinstance(vehicles, list):
            fail("RTA input", ["RTA JSON is not an array"])
        rta = verify_rta(master, vehicles)
        print(f"rta_verification=PASS matched={rta['matched']} ambiguous={rta['ambiguous']} "
              f"unmatched={rta['unmatched']} ymm_match={rta['ymm_match']} "
              f"ymm_mismatch={rta['ymm_mismatch']} vin_exact={rta['vin_exact']} "
              f"vin_conflict={rta['vin_conflict']}")
        print(f"  ymm_mismatch_assets={rta['ymm_mismatch_assets']}")

    # --- 5. Build entries + completeness cross-check (AC#7) ---
    entries = build_entries(master, tire, ymm_map, classes)
    completeness = check_completeness(entries)
    print(f"completeness=PASS complete={completeness['complete']} partial={completeness['partial']} "
          f"blank={completeness['blank']} detail_populated={completeness['detail_populated']}")

    # --- 6. Assemble manifest ---
    manifest = {
        "schema": "tire-reconciliation-manifest/1",
        "issue": 16,
        "generated_at": dt.datetime.now().astimezone().isoformat(timespec="seconds"),
        "sources": {
            "master_path": str(master_path),
            "master_sha256": master_sha,
            "tire_path": str(tire_path),
            "tire_sha256": tire_sha,
            "output_path": str(output_path),
            "output_absent": True,
        },
        "column_map": {
            "master": {name: mcols[name] for name in mcols},
            "master_blank_column": 12,
            "appended_fields": {
                name: APPENDED_START_COL + i for i, name in enumerate(APPENDED_FIELDS)
            },
            "fill_directives": {FILL_YELLOW: "force yellow on Front+Rear tire cells (hex resolved in #17)"},
        },
        "rta_verification": rta,
        "category_totals": {
            "direct": len(classes["direct"]),
            "moved_detail": len(classes["moved_detail"]),
            "corrected": len(classes["corrected"]),
            "yellow": len(classes["yellow"]),
            "unresolved": len(classes["unresolved"]),
        },
        "category_members": {
            name: sorted(members) for name, members in classes.items()
        },
        "tire_completeness": {
            "complete": completeness["complete"],
            "partial": completeness["partial"],
            "blank": completeness["blank"],
            "detail_populated": completeness["detail_populated"],
            "partial_assets": completeness["partial_assets"],
        },
        "entries": entries,
    }

    if args.dry_run:
        print("\nDRY RUN complete — manifest not written.")
        return
    if args.skip_rta:
        manifest["_warning"] = "RTA verification SKIPPED; not valid for #17 consumption."

    manifest_path.parent.mkdir(parents=True, exist_ok=True)
    manifest_path.write_text(json.dumps(manifest, indent=2, sort_keys=False) + "\n")
    manifest_sha = sha256(manifest_path)
    print(f"\n=== MANIFEST WRITTEN ===")
    print(f"  path={manifest_path}")
    print(f"  entries={len(entries)}")
    print(f"  sha256={manifest_sha}")


if __name__ == "__main__":
    main()
